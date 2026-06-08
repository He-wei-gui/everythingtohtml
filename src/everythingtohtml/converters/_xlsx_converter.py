"""XLSX -> HTML, one ``<table>`` per worksheet (requires the ``xlsx`` extra)."""

from __future__ import annotations

from typing import Any, BinaryIO

from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MissingDependencyException
from .._html_builder import escape_text, wrap_document
from .._stream_info import StreamInfo

__all__ = ["XlsxConverter"]

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class XlsxConverter(DocumentConverter):
    """Convert Excel ``.xlsx`` workbooks into one HTML table per sheet."""

    priority = DocumentConverter.PRIORITY_SPECIFIC_FILE_FORMAT

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        ext = stream_info.normalized_extension()
        mimetype = (stream_info.mimetype or "").split(";", 1)[0].strip().lower()
        return ext in (".xlsx", ".xlsm") or mimetype == _XLSX_MIME

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise MissingDependencyException(
                "XLSX support requires the 'xlsx' extra. "
                "Install it with: pip install everythingtohtml[xlsx]"
            ) from exc

        workbook = load_workbook(file_stream, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in workbook.worksheets:
            parts.append(f"<section><h2>{escape_text(sheet.title)}</h2>")
            parts.append(_sheet_to_table(sheet))
            parts.append("</section>")
        workbook.close()

        if not parts:
            parts.append("<p><em>(empty workbook)</em></p>")

        title = stream_info.filename
        html = wrap_document("\n".join(parts), title=title)
        return DocumentConverterResult(html, title=title)


def _sheet_to_table(sheet: Any) -> str:
    rows = list(sheet.iter_rows(values_only=True))
    # Trim trailing fully-empty rows that openpyxl sometimes reports.
    while rows and all(cell is None for cell in rows[-1]):
        rows.pop()
    if not rows:
        return "<p><em>(empty sheet)</em></p>"

    parts = ["<table>"]
    header, *body = rows
    parts.append("<thead><tr>")
    parts.extend(f"<th>{escape_text(_fmt(cell))}</th>" for cell in header)
    parts.append("</tr></thead><tbody>")
    for row in body:
        parts.append("<tr>")
        parts.extend(f"<td>{escape_text(_fmt(cell))}</td>" for cell in row)
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


def _fmt(cell: Any) -> str:
    return "" if cell is None else str(cell)
